import re
import time
import openpyxl
from openpyxl.styles import PatternFill, Font
from tqdm import tqdm
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

import config
from modules.glpi_base import (
    login, selecionar_planilha, localizar_cabecalho,
    salvar_output, limpar_ticket_id,
)

FILL_ERRO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FONT_ERRO = Font(color="FF0000", bold=True)

SELETORES_CLOSE = [
    'input[name="cleartimefield"]',
    'input[name="closedate"]',
    'input[id="closedate"]',
    'input[name="time_solved"]',
    'input[id="time_solved"]',
]
PADROES_INICIO = [
    re.compile(r"in[ií]cio\s+de\s+atendimento[:\s]+(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2})", re.IGNORECASE),
    re.compile(r"in[ií]cio\s+de\s+atendimento[:\s]+(\d{2}/\d{2}/\d{4}[ T]\d{2}:\d{2})", re.IGNORECASE),
]
DATA_RE = re.compile(
    r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}"
    r"|^\d{2}/\d{2}/\d{4}[ T]\d{2}:\d{2}"
)


def _e_data(valor: str) -> bool:
    return bool(DATA_RE.match(valor.strip()))


def _marcar_erro(cell, texto):
    cell.value = texto
    cell.fill  = FILL_ERRO
    cell.font  = FONT_ERRO


def scrape_chamado(page, ticket_id: int) -> dict:
    url = f"{config.GLPI_URL}/front/ticket.form.php?id={ticket_id}"
    page.goto(url, wait_until="domcontentloaded", timeout=30_000)

    try:
        page.wait_for_selector(
            'select[data-actor-type="assign"] option[selected="true"]',
            timeout=10_000,
        )
    except PWTimeout:
        try:
            page.wait_for_load_state("networkidle", timeout=10_000)
        except PWTimeout:
            pass

    resultado = {"analista": None, "inicio_atendimento": None, "final_atendimento": None}
    html = page.content()
    analistas_lower = {a.lower(): a for a in config.ANALISTAS}

    # ── Analista A: Assigned To ──
    for opt in page.query_selector_all(
        'select[data-actor-type="assign"] option[selected="true"][data-itemtype="User"]'
    ):
        nome = (opt.inner_text() or "").strip()
        if analistas_lower.get(nome.lower()):
            resultado["analista"] = analistas_lower[nome.lower()]
            break

    # ── Analista B: links no timeline da solução ──
    if not resultado["analista"]:
        for link in page.query_selector_all('a[href*="/front/user.form.php"][title]'):
            nome = (link.get_attribute("title") or "").strip()
            if analistas_lower.get(nome.lower()):
                resultado["analista"] = analistas_lower[nome.lower()]
                break

    # ── Analista C: tag NCCL ──
    if not resultado["analista"]:
        for chip in page.query_selector_all("li.select2-selection__choice"):
            texto = (chip.get_attribute("title") or chip.inner_text() or "").strip()
            if "nccl" in texto.lower():
                resultado["analista"] = "NCCL"
                break
        if not resultado["analista"]:
            for opt in page.query_selector_all('select[name="_plugin_tag_tag_values[]"] option'):
                texto = (opt.inner_text() or opt.get_attribute("data-title") or "").strip()
                if "nccl" in texto.lower():
                    resultado["analista"] = "NCCL"
                    break

    # ── Início do atendimento ──
    for padrao in PADROES_INICIO:
        matches = padrao.findall(html)
        if matches:
            resultado["inicio_atendimento"] = matches[-1]
            break

    # ── Final do atendimento ──
    for seletor in SELETORES_CLOSE:
        try:
            elem = page.query_selector(seletor)
            if elem:
                valor = (elem.get_attribute("value") or "").strip()
                if valor and _e_data(valor):
                    resultado["final_atendimento"] = valor
                    break
        except Exception:
            continue

    if not resultado["final_atendimento"]:
        for campo in ("cleartimefield", "time_solved", "closedate"):
            m = re.search(
                rf'name=["\'{campo}["\'][^>]*value=["\']([^"\']+)["\']'
                rf'|value=["\']([^"\']+)["\'][^>]*name=["\'{campo}["\']',
                html, re.IGNORECASE,
            )
            if m:
                valor = (m.group(1) or m.group(2) or "").strip()
                if _e_data(valor):
                    resultado["final_atendimento"] = valor
                    break

    return resultado


def run():
    print("\n─── Scrapping de Chamados ───────────────────────────────")

    caminho = selecionar_planilha()
    if not caminho:
        return False

    print(f"\n  [EXCEL] Abrindo '{caminho.name}' ...")
    try:
        wb = openpyxl.load_workbook(caminho)
    except Exception as e:
        print(f"  [ERRO] Não foi possível abrir o arquivo: {e}")
        return

    ws = wb.active
    NOMES_COLUNA_TICKET = [config.COLUNA_TICKET, "ID", "TICKET", "Ticket", "Id", "id", "glpi"]
    linha_cabecalho, cabecalho = None, {}
    for nome in NOMES_COLUNA_TICKET:
        linha_cabecalho, cabecalho = localizar_cabecalho(ws, nome)
        if linha_cabecalho:
            config.COLUNA_TICKET = nome   # usa o nome encontrado nas próximas referências
            break

    if not linha_cabecalho:
        print(f"  [ERRO] Nenhuma coluna de ticket encontrada nas primeiras 10 linhas.")
        print(f"         Esperado um destes nomes: {NOMES_COLUNA_TICKET}")
        return

    print(f"  [EXCEL] Cabeçalho encontrado na linha {linha_cabecalho}")

    col_max = ws.max_column
    for col_nome in [config.COLUNA_ANALISTA, config.COLUNA_INICIO_ATENDIMENTO, config.COLUNA_FINAL_ATENDIMENTO]:
        if col_nome not in cabecalho:
            col_max += 1
            ws.cell(row=linha_cabecalho, column=col_max, value=col_nome)
            cabecalho[col_nome] = col_max
            print(f"  [EXCEL] Coluna '{col_nome}' criada na posição {col_max}")

    col_ticket = cabecalho[config.COLUNA_TICKET]
    col_anal   = cabecalho[config.COLUNA_ANALISTA]
    col_inicio = cabecalho[config.COLUNA_INICIO_ATENDIMENTO]
    col_final  = cabecalho[config.COLUNA_FINAL_ATENDIMENTO]

    def _celula_vazia(cell):
        return not str(cell.value or "").strip()

    linhas = [
        row[col_ticket - 1].row
        for row in ws.iter_rows(min_row=linha_cabecalho + 1)
        if not _celula_vazia(row[col_ticket - 1]) and _celula_vazia(row[col_anal - 1])
    ]

    total = len(linhas)
    print(f"  [EXCEL] {total} chamados para processar\n")
    if total == 0:
        print("  [OK] Nada para processar.")
        return

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context()
        page    = context.new_page()

        page.route(
            "**/*",
            lambda route: route.abort()
            if route.request.resource_type in {"image", "media", "font", "stylesheet"}
            else route.continue_(),
        )

        try:
            login(page)
        except Exception as e:
            print(f"  [ERRO] Falha no login: {e}")
            browser.close()
            return

        erros = []

        for row_num in tqdm(linhas, desc="  Processando", unit="chamado"):
            ticket_id = ws.cell(row=row_num, column=col_ticket).value
            ticket_num = limpar_ticket_id(ticket_id)

            if not ticket_num:
                tqdm.write(f"  [AVISO] Ticket inválido ignorado: '{ticket_id}'")
                continue

            sucesso = False
            for tentativa in range(1, config.MAX_TENTATIVAS + 1):
                try:
                    dados = scrape_chamado(page, ticket_num)

                    cell_anal = ws.cell(row=row_num, column=col_anal)
                    if dados["analista"]:
                        cell_anal.value = dados["analista"]
                    else:
                        _marcar_erro(cell_anal, "NÃO ENCONTRADO")

                    cell_inicio = ws.cell(row=row_num, column=col_inicio)
                    if dados["inicio_atendimento"]:
                        cell_inicio.value = dados["inicio_atendimento"]
                    else:
                        _marcar_erro(cell_inicio, "NÃO ENCONTRADO")

                    cell_final = ws.cell(row=row_num, column=col_final)
                    if dados["final_atendimento"]:
                        cell_final.value = dados["final_atendimento"]
                    else:
                        _marcar_erro(cell_final, "NÃO ENCONTRADO")

                    sucesso = True
                    break

                except PWTimeout:
                    tqdm.write(f"  [TIMEOUT] Ticket {ticket_id} — tentativa {tentativa}/{config.MAX_TENTATIVAS}")
                except Exception as exc:
                    tqdm.write(f"  [ERRO] Ticket {ticket_id}: {exc} — tentativa {tentativa}/{config.MAX_TENTATIVAS}")
                time.sleep(2)

            if not sucesso:
                erros.append(ticket_id)

            time.sleep(config.DELAY_ENTRE_CHAMADOS)

        browser.close()

    destino = salvar_output(wb, caminho)
    print(f"\n  [OK] Planilha salva em 'output/{destino.name}'")

    if erros:
        print(f"\n  [ATENÇÃO] {len(erros)} chamados falharam após {config.MAX_TENTATIVAS} tentativas:")
        for t in erros:
            print(f"    Ticket #{t}")
