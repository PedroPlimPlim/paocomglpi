import time
import openpyxl
from tqdm import tqdm
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

import config
from modules.glpi_base import (
    login, selecionar_planilha, localizar_cabecalho,
    salvar_output, limpar_ticket_id,
)

# Colunas alvo no Excel
COLUNA_ANALISTA   = "Analista"
COLUNA_RESULTADO  = "Resultado"
COLUNA_OBSERVACAO = "Observação"

# Possíveis nomes da coluna de ticket na planilha
NOMES_COLUNA_TICKET = ["GLPI", "ID", "TICKET", "Ticket", "Id", "id", "glpi"]

# Critérios que classificam o chamado como Elsys (busca no HTML, case-insensitive)
CRITERIOS_ELSYS = [
    "TECNOLOGIA: ELSYS",
    "OPERADORA: TIM",
    "TOPOLOGIA: 2",
    "TIPO DE SITE: 2",
]


def _e_elsys(html: str) -> bool:
    html_lower = html.lower()
    return any(c.lower() in html_lower for c in CRITERIOS_ELSYS)


def _encontrar_coluna_ticket(cabecalho: dict) -> int | None:
    for nome in NOMES_COLUNA_TICKET:
        if nome in cabecalho:
            return cabecalho[nome]
    return None


def run():
    print("\n─── Café com Elsys ──────────────────────────────────────")

    # Pergunta o nome do analista que está rodando
    nome_analista = ""
    while not nome_analista:
        nome_analista = input("\n  Seu nome (será preenchido na coluna Analista): ").strip()
        if not nome_analista:
            print("  Nome não pode estar vazio.")

    caminho = selecionar_planilha()
    if not caminho:
        return False

    print(f"\n  [EXCEL] Abrindo '{caminho.name}' ...")
    try:
        wb = openpyxl.load_workbook(caminho)
    except Exception as e:
        print(f"  [ERRO] Não foi possível abrir o arquivo: {e}")
        return

    ws = wb.active

    # Localiza cabeçalho — testa cada possível nome de coluna de ticket
    linha_cabecalho = None
    cabecalho = {}
    for nome_ticket in NOMES_COLUNA_TICKET:
        linha_cabecalho, cabecalho = localizar_cabecalho(ws, nome_ticket)
        if linha_cabecalho:
            break

    if not linha_cabecalho:
        print(f"  [ERRO] Nenhuma coluna de ticket encontrada.")
        print(f"         Esperado um destes nomes: {NOMES_COLUNA_TICKET}")
        return

    print(f"  [EXCEL] Cabeçalho encontrado na linha {linha_cabecalho}")

    col_ticket = _encontrar_coluna_ticket(cabecalho)

    # Cria colunas de saída se não existirem
    col_max = ws.max_column
    for col_nome in [COLUNA_ANALISTA, COLUNA_RESULTADO, COLUNA_OBSERVACAO]:
        if col_nome not in cabecalho:
            col_max += 1
            ws.cell(row=linha_cabecalho, column=col_max, value=col_nome)
            cabecalho[col_nome] = col_max
            print(f"  [EXCEL] Coluna '{col_nome}' criada na posição {col_max}")

    col_anal  = cabecalho[COLUNA_ANALISTA]
    col_res   = cabecalho[COLUNA_RESULTADO]
    col_obs   = cabecalho[COLUNA_OBSERVACAO]

    # Linhas pendentes: tem ticket e analista ainda não preenchido
    linhas = [
        row[col_ticket - 1].row
        for row in ws.iter_rows(min_row=linha_cabecalho + 1)
        if row[col_ticket - 1].value and not row[col_anal - 1].value
    ]

    total = len(linhas)
    print(f"  [EXCEL] {total} chamados para processar\n")
    if total == 0:
        print("  [OK] Nada para processar.")
        return

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context()
        page    = context.new_page()

        page.route(
            "**/*",
            lambda route: route.abort()
            if route.request.resource_type in {"image", "media", "font", "stylesheet"}
            else route.continue_(),
        )

        try:
            login(page)
        except Exception as e:
            print(f"  [ERRO] Falha no login: {e}")
            browser.close()
            return

        erros = []

        for row_num in tqdm(linhas, desc="  Processando", unit="chamado"):
            ticket_raw = ws.cell(row=row_num, column=col_ticket).value
            ticket_id  = limpar_ticket_id(ticket_raw)

            if not ticket_id:
                tqdm.write(f"  [AVISO] Ticket inválido ignorado: '{ticket_raw}'")
                continue

            sucesso = False
            for tentativa in range(1, config.MAX_TENTATIVAS + 1):
                try:
                    url = f"{config.GLPI_URL}/front/ticket.form.php?id={ticket_id}"
                    page.goto(url, wait_until="domcontentloaded", timeout=30_000)
                    try:
                        page.wait_for_load_state("networkidle", timeout=8_000)
                    except PWTimeout:
                        pass

                    html = page.content()

                    if _e_elsys(html):
                        ws.cell(row=row_num, column=col_anal).value = nome_analista
                        ws.cell(row=row_num, column=col_res).value  = "Elsys"
                        ws.cell(row=row_num, column=col_obs).value  = "Elsys"

                    sucesso = True
                    break

                except PWTimeout:
                    tqdm.write(f"  [TIMEOUT] Ticket {ticket_id} — tentativa {tentativa}/{config.MAX_TENTATIVAS}")
                except Exception as exc:
                    tqdm.write(f"  [ERRO] Ticket {ticket_id}: {exc} — tentativa {tentativa}/{config.MAX_TENTATIVAS}")
                time.sleep(2)

            if not sucesso:
                erros.append(ticket_id)

            time.sleep(config.DELAY_ENTRE_CHAMADOS)

        browser.close()

    destino = salvar_output(wb, caminho)
    print(f"\n  [OK] Planilha salva em 'output/{destino.name}'")

    if erros:
        print(f"\n  [ATENÇÃO] {len(erros)} chamados falharam após {config.MAX_TENTATIVAS} tentativas:")
        for t in erros:
            print(f"    Ticket #{t}")
